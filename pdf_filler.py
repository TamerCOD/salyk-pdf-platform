# -*- coding: utf-8 -*-
"""
Заполнение PDF-формы ZVN STI - 222 данными из Excel.
Адаптировано для использования как модуль из веб-приложения.
"""

import io
import os
import re
import sys
from pathlib import Path

import openpyxl
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

PAGE_W, PAGE_H = 595.2, 841.8  # A4

FONT_NAME = 'AppFont'
FONT_BOLD = 'AppFont-Bold'

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_FONT_CANDIDATES = [
    ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
     '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
    ('/usr/share/fonts/dejavu/DejaVuSans.ttf',
     '/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf'),
    ('C:/Windows/Fonts/arial.ttf',
     'C:/Windows/Fonts/arialbd.ttf'),
    ('C:/Windows/Fonts/ARIAL.TTF',
     'C:/Windows/Fonts/ARIALBD.TTF'),
    ('/System/Library/Fonts/Supplemental/Arial.ttf',
     '/System/Library/Fonts/Supplemental/Arial Bold.ttf'),
    (os.path.join(_SCRIPT_DIR, 'assets', 'DejaVuSans.ttf'),
     os.path.join(_SCRIPT_DIR, 'assets', 'DejaVuSans-Bold.ttf')),
    (os.path.join(_SCRIPT_DIR, 'DejaVuSans.ttf'),
     os.path.join(_SCRIPT_DIR, 'DejaVuSans-Bold.ttf')),
]

_font_loaded = False
for reg, bold in _FONT_CANDIDATES:
    if os.path.exists(reg):
        try:
            pdfmetrics.registerFont(TTFont(FONT_NAME, reg))
            if os.path.exists(bold):
                pdfmetrics.registerFont(TTFont(FONT_BOLD, bold))
            else:
                FONT_BOLD = FONT_NAME
            _font_loaded = True
            break
        except Exception:
            continue

if not _font_loaded:
    print('WARNING: cyrillic TTF font not found, falling back to Helvetica', file=sys.stderr)
    FONT_NAME = 'Helvetica'
    FONT_BOLD = 'Helvetica-Bold'


CELL_GROUPS = {
    'inn_taxpayer': {
        'centers': [188.75, 206.25, 223.75, 241.3, 258.85, 276.35, 293.85,
                    311.35, 328.85, 346.35, 363.9, 381.4, 398.95, 416.45],
        'top': 121.9, 'height': 17.8,
    },
    'org_code': {
        'centers': [189.0, 206.9, 224.9],
        'top': 178.2, 'height': 17.8,
    },
    'date_spravka': {
        'centers': [424.25, 437.85, 462.05, 477.15, 499.85, 514.95, 530.05, 545.2],
        'top': 225.4, 'height': 17.0,
    },
    'date_protokol': {
        'centers': [425.3, 438.8, 462.8, 477.8, 500.3, 515.3, 530.3, 545.3],
        'top': 256.5, 'height': 17.0,
    },
    'date_decision': {
        'centers': [426.0, 439.5, 463.5, 478.5, 501.0, 516.0, 531.0, 546.0],
        'top': 534.4, 'height': 17.0,
    },
    'inn_employee': {
        'centers': [229.65, 243.9, 258.15, 272.45, 286.75, 301.0, 315.25,
                    329.55, 343.85, 358.15, 372.4, 386.65, 400.95, 415.25],
        'top': 646.0, 'height': 17.6,
    },
    'date_delivery': {
        'centers': [209.3, 222.8, 246.8, 261.8, 284.3, 299.3, 314.3, 329.3],
        'top': 773.9, 'height': 16.9,
    },
}

TEXT_FIELDS = {
    'taxpayer_name':   {'x0': 180.3, 'x1': 552.8, 'top': 148.4, 'h': 17.8},
    'org_name':        {'x0': 240.9, 'x1': 552.8, 'top': 178.2, 'h': 17.8},
    'spravka_num':     {'x0': 155.9, 'x1': 290.9, 'top': 225.3, 'h': 18.0},
    'protokol_num':    {'x0': 155.9, 'x1': 290.9, 'top': 256.5, 'h': 18.0},
    'decision_num':    {'x0': 160.2, 'x1': 295.2, 'top': 536.1, 'h': 18.0},
    'head_name':       {'x0':  85.0, 'x1': 481.0, 'top': 567.6, 'h': 24.7},
    'chief_name':      {'x0':  85.5, 'x1': 481.5, 'top': 604.3, 'h': 24.6},
    'employee_name':   {'x0':  85.9, 'x1': 481.9, 'top': 683.9, 'h': 24.8},
    'taxpayer_signer': {'x0':  85.9, 'x1': 481.9, 'top': 739.8, 'h': 24.8},
}

TABLE_ROW = {
    'top': 473.4, 'h': 28.3,
    'cols': [
        {'name': 'kbk',     'x0':  34.0, 'x1': 170.1},
        {'name': 'pay_name','x0': 170.1, 'x1': 311.8},
        {'name': 'pay_date','x0': 311.8, 'x1': 453.5},
        {'name': 'pay_sum', 'x0': 453.5, 'x1': 566.2},
    ],
}

# Соответствие букв колонок → название поля для UI группировки
COLUMN_LABELS = {
    'A': '№',
    'B': 'ИНН налогоплательщика',
    'C': 'ФИО/наименование налогоплательщика',
    'D': 'Код налогового органа',
    'E': 'Наименование налогового органа',
    'F': '№ справки обследования',
    'G': 'Дата справки обследования',
    'H': 'Номер протокола комиссии',
    'I': 'Дата протокола комиссии',
    'J': 'Номер принятого Решения',
    'K': 'Дата принятия Решения',
    'L': 'ФИО руководителя налогового органа',
    'M': 'ФИО начальника отдела',
    'N': 'ИНН сотрудника налогового органа',
    'O': 'ФИО сотрудника налогового органа',
    'P': 'ФИО налогоплательщика (расписка)',
    'Q': 'Дата вручения Решения',
    'R': 'Код бюджетной классификации',
    'S': 'Наименование неналогового платежа',
    'T': 'Дата возникновения платежей',
    'U': 'Сумма неналогового платежа',
}


def _draw_in_cells(c, group_key, value, font_size=11):
    g = CELL_GROUPS[group_key]
    if not value:
        return
    s = re.sub(r'[^\d]', '', str(value))
    centers = g['centers']
    y_center_top = g['top'] + g['height'] / 2
    y_baseline = PAGE_H - y_center_top - font_size * 0.35
    c.setFont(FONT_NAME, font_size)
    for ch, cx in zip(s, centers):
        c.drawCentredString(cx, y_baseline, ch)


def _truncate_to_width(c, text, font, size, max_width):
    if not text:
        return '', size
    t = str(text)
    cur = size
    while pdfmetrics.stringWidth(t, font, cur) > max_width and cur > 6:
        cur -= 0.5
    if pdfmetrics.stringWidth(t, font, cur) > max_width:
        while t and pdfmetrics.stringWidth(t + '…', font, cur) > max_width:
            t = t[:-1]
        t += '…'
    return t, cur


def _wrap_text(text, font, size, max_width):
    if not text:
        return []
    words = str(text).split()
    lines = []
    current = ''
    for word in words:
        candidate = (current + ' ' + word).strip()
        if pdfmetrics.stringWidth(candidate, font, size) <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            if pdfmetrics.stringWidth(word, font, size) > max_width:
                chunk = ''
                for ch in word:
                    if pdfmetrics.stringWidth(chunk + ch, font, size) <= max_width:
                        chunk += ch
                    else:
                        lines.append(chunk)
                        chunk = ch
                current = chunk
            else:
                current = word
    if current:
        lines.append(current)
    return lines


def _fit_text_in_box(text, font, base_size, max_width, max_lines=2, min_size=6.5):
    if not text:
        return [], base_size
    text = str(text).strip()
    cur = base_size
    while cur >= min_size:
        if pdfmetrics.stringWidth(text, font, cur) <= max_width:
            return [text], cur
        cur -= 0.5
    for n_lines in range(2, max_lines + 1):
        cur = base_size
        while cur >= min_size:
            lines = _wrap_text(text, font, cur, max_width)
            if len(lines) <= n_lines:
                return lines, cur
            cur -= 0.5
    lines = _wrap_text(text, font, min_size, max_width)
    if len(lines) > max_lines:
        kept = lines[:max_lines]
        last = kept[-1]
        while last and pdfmetrics.stringWidth(last + '…', font, min_size) > max_width:
            last = last[:-1]
        kept[-1] = last + '…'
        return kept, min_size
    return lines, min_size


def _draw_text_in_field(c, field_key, value, font_size=10, align='left', pad=3):
    if value is None or value == '':
        return
    f = TEXT_FIELDS[field_key]
    max_w = f['x1'] - f['x0'] - 2 * pad
    text, size = _truncate_to_width(c, str(value), FONT_NAME, font_size, max_w)
    bottom_offset = 4 if f['h'] < 20 else 9
    y_baseline = PAGE_H - (f['top'] + f['h']) + bottom_offset
    c.setFont(FONT_NAME, size)
    if align == 'center':
        c.drawCentredString((f['x0'] + f['x1']) / 2, y_baseline, text)
    else:
        c.drawString(f['x0'] + pad, y_baseline, text)


def _format_date(v):
    if v is None or v == '':
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%d%m%Y')
    s = str(v).strip()
    digits = re.sub(r'[^\d]', '', s)
    return digits


def _format_date_text(v):
    if v is None or v == '':
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%d.%m.%Y')
    s = str(v).strip()
    digits = re.sub(r'[^\d]', '', s)
    if len(digits) == 8:
        return f'{digits[:2]}.{digits[2:4]}.{digits[4:]}'
    return s


def _draw_table_row(c, row_data, base_size=9):
    tr = TABLE_ROW
    cell_top = tr['top']
    cell_bottom = tr['top'] + tr['h']
    cell_mid_y_reportlab = PAGE_H - (cell_top + cell_bottom) / 2

    for col in tr['cols']:
        name = col['name']
        value = row_data.get(name, '')
        if value is None or value == '':
            continue
        max_w = col['x1'] - col['x0'] - 6
        lines, size = _fit_text_in_box(value, FONT_NAME, base_size, max_w,
                                       max_lines=2, min_size=6.5)
        line_h = size * 1.15
        c.setFont(FONT_NAME, size)
        x_center = (col['x0'] + col['x1']) / 2
        total_h = line_h * len(lines)
        first_baseline = cell_mid_y_reportlab + total_h / 2 - size * 0.85
        for i, line in enumerate(lines):
            y = first_baseline - i * line_h
            c.drawCentredString(x_center, y, line)


def build_overlay(data):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(PAGE_W, PAGE_H))

    _draw_in_cells(c, 'inn_taxpayer', data.get('B'))
    _draw_text_in_field(c, 'taxpayer_name', data.get('C'))
    _draw_in_cells(c, 'org_code', data.get('D'))
    _draw_text_in_field(c, 'org_name', data.get('E'))

    _draw_text_in_field(c, 'spravka_num', data.get('F'))
    _draw_in_cells(c, 'date_spravka', _format_date(data.get('G')))
    _draw_text_in_field(c, 'protokol_num', data.get('H'))
    _draw_in_cells(c, 'date_protokol', _format_date(data.get('I')))

    _draw_table_row(c, {
        'kbk':      data.get('R', ''),
        'pay_name': data.get('S', ''),
        'pay_date': _format_date_text(data.get('T', '')),
        'pay_sum':  data.get('U', ''),
    })

    _draw_text_in_field(c, 'decision_num', data.get('J'))
    _draw_in_cells(c, 'date_decision', _format_date(data.get('K')))
    _draw_text_in_field(c, 'head_name', data.get('L'))
    _draw_text_in_field(c, 'chief_name', data.get('M'))
    _draw_in_cells(c, 'inn_employee', data.get('N'))
    _draw_text_in_field(c, 'employee_name', data.get('O'))

    _draw_text_in_field(c, 'taxpayer_signer', data.get('P'))
    _draw_in_cells(c, 'date_delivery', _format_date(data.get('Q')))

    c.save()
    buf.seek(0)
    return buf


def fill_pdf_bytes(template_path, data):
    """Возвращает bytes заполненного PDF (одна страница)."""
    overlay_buf = build_overlay(data)
    template = PdfReader(template_path)
    overlay = PdfReader(overlay_buf)

    writer = PdfWriter()
    page = template.pages[0]
    page.merge_page(overlay.pages[0])
    writer.add_page(page)

    out = io.BytesIO()
    writer.write(out)
    out.seek(0)
    return out.getvalue()


def merge_pdfs(pdf_bytes_list):
    """Склеивает список pdf-байтов в один многостраничный PDF."""
    writer = PdfWriter()
    for pdf_bytes in pdf_bytes_list:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        for page in reader.pages:
            writer.add_page(page)
    out = io.BytesIO()
    writer.write(out)
    out.seek(0)
    return out.getvalue()


def read_excel_rows(xlsx_path_or_bytes):
    """Читает Excel и возвращает список dict {колонка_буква: значение}."""
    if isinstance(xlsx_path_or_bytes, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_path_or_bytes), data_only=True)
    else:
        wb = openpyxl.load_workbook(xlsx_path_or_bytes, data_only=True)
    ws = wb.active
    rows = []
    cols = list('ABCDEFGHIJKLMNOPQRSTU')
    for r in range(2, ws.max_row + 1):
        row = {}
        for col in cols:
            v = ws[f'{col}{r}'].value
            row[col] = v
        meaningful = any(row.get(c) not in (None, '') for c in 'BCDEFGHIJKLMNOPQRSTU')
        if meaningful:
            rows.append(row)
    return rows


def safe_filename(s):
    s = re.sub(r'[\\/:*?"<>|\n\r\t]', '_', str(s))
    return s.strip().rstrip('.')[:80] or 'file'
