# -*- coding: utf-8 -*-
"""
Получение наименований по ИНН — cabinet.salyk.kg
Адаптировано для использования как модуль из веб-приложения.
"""

import io
import time
from typing import Iterable

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


API_URL = "https://cabinet.salyk.kg/TinCheck/GetTaxPayer"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer":         "https://cabinet.salyk.kg/account/register",
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "ru-RU,ru;q=0.9",
}
DEFAULT_DELAY = 0.4


def _new_session():
    s = requests.Session()
    s.headers.update(HEADERS)
    try:
        s.get("https://cabinet.salyk.kg/account/register", timeout=10)
    except Exception:
        pass
    return s


def lookup_inn(session, inn):
    inn = str(inn).strip()
    result = {"inn": inn, "name": "", "rayon": "", "status": "ok"}
    try:
        resp = session.get(API_URL, params={"Tin": inn}, timeout=15)
        if resp.status_code == 400:
            result["status"] = "не найдено"
            return result
        if resp.status_code != 200:
            result["status"] = f"HTTP {resp.status_code}"
            return result
        text = resp.text.strip()
        if not text or text == "null":
            result["status"] = "не найдено"
            return result
        data = resp.json()
        if data is None:
            result["status"] = "не найдено"
        elif isinstance(data, dict):
            result["name"]  = data.get("name", "")
            result["rayon"] = data.get("rayon", "")
            if not result["name"]:
                result["status"] = "не найдено"
        else:
            result["status"] = "не найдено"
    except requests.exceptions.Timeout:
        result["status"] = "таймаут"
    except requests.exceptions.ConnectionError:
        result["status"] = "нет соединения"
    except Exception as e:
        result["status"] = "ошибка"
    return result


def lookup_many(inns: Iterable[str], delay: float = DEFAULT_DELAY, progress=None):
    session = _new_session()
    results = []
    inns = list(inns)
    for i, inn in enumerate(inns, 1):
        r = lookup_inn(session, inn)
        results.append(r)
        if progress:
            progress(i, len(inns), r)
        if i < len(inns):
            time.sleep(delay)
    return results


def parse_inn_list_from_excel(xlsx_bytes):
    """Берёт ИНН из первой колонки Excel-файла."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    inns = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        v = row[0]
        if v is None:
            continue
        s = str(v).strip()
        if s.lower() in ("инн", "inn", "иин", "тин", "tin"):
            continue
        if s.isdigit() and len(s) >= 10:
            inns.append(s)
    wb.close()
    return inns


def parse_inn_list_from_text(text):
    out = []
    for token in (text or "").replace(",", "\n").replace(";", "\n").split():
        t = token.strip()
        if t.isdigit() and len(t) >= 10:
            out.append(t)
    return out


def build_results_xlsx(results):
    """Возвращает bytes Excel с результатами."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"

    thin     = Side(style="thin", color="BFBFBF")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    ok_fill  = PatternFill("solid", fgColor="E2EFDA")
    warn     = PatternFill("solid", fgColor="FFF2CC")
    err      = PatternFill("solid", fgColor="FCE4D6")

    headers = ["ИНН", "Наименование / ФИО", "Код УГНС", "Статус"]
    widths  = [18, 60, 28, 16]
    ws.row_dimensions[1].height = 22
    for col, (hdr, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    for i, r in enumerate(results, 2):
        values = [r["inn"], r["name"], r["rayon"], r["status"]]
        if r["name"]:
            row_fill = ok_fill
        elif r["status"] == "не найдено":
            row_fill = warn
        else:
            row_fill = err
        ws.row_dimensions[i].height = 18
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = row_fill; cell.border = border
            cell.alignment = center if col in (1, 3, 4) else left_al

    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
